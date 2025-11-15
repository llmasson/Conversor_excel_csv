import openpyxl
import csv

def converter_xlsx_para_csv(caminho_xlsx, caminho_csv, nome_planilha=None):
    wb = openpyxl.load_workbook(caminho_xlsx, read_only=True, data_only=True)

    print("Planilhas disponíveis:", wb.sheetnames)
    ws = wb[nome_planilha] if nome_planilha else wb.active
    print("Usando a planilha:", ws.title)

    contador = 0
    with open(caminho_csv, mode='w', newline='', encoding='utf-8') as arquivo_csv:
        escritor = csv.writer(arquivo_csv, delimiter=';')

        for linha in ws.iter_rows(values_only=True):
            linha_limpa = [cel if cel is not None else '' for cel in linha]
            if any(linha_limpa):  # Ignora linhas completamente vazias
                escritor.writerow(linha_limpa)
                contador += 1

    wb.close()

    if contador == 0:
        print("⚠️ Nenhuma linha foi escrita. A planilha pode estar vazia ou o nome da aba pode estar incorreto.")
    else:
        print(f'✅ Conversão concluída: {contador} linhas salvas em "{caminho_csv}"')

# 🧪 Exemplo de uso:
#converter_xlsx_para_csv('C:\Pasta_trabalho\GitLab\Python\Arquivamento Negócios Pipeline Legado (Parcial).xlsx', 'C:\Pasta_trabalho\GitLab\Python\saida.csv')


def converter_excel():
    import pandas as pd

    arquivo = r'C:\Pasta_trabalho\GitLab\Python\Arquivamento Negócios Pipeline Legado (Parcial).xlsx'
    aba = 'Todos os negócios'
    saida_csv = r'C:\Pasta_trabalho\GitLab\Python\saida_negocios.csv'

    try:
        print(f"Lendo e limpando aba: {aba}")
        df = pd.read_excel(arquivo, sheet_name=aba, engine='openpyxl')

        # Remove linhas e colunas totalmente vazias
        df.dropna(how='all', inplace=True)
        df.dropna(axis=1, how='all', inplace=True)

        print(f"Linhas válidas: {len(df)}, Colunas: {len(df.columns)}")

        # Salva o CSV
        df.to_csv(saida_csv, sep=';', index=False, encoding='utf-8')
        print(f"✅ CSV salvo com sucesso em: {saida_csv}")

    except Exception as e:
        print("❌ Erro ao converter:", e)

converter_excel()

import pandas as pd
import os

def dividir_em_csvs(df, pasta_saida, tamanho_lote=5000, nome_base='parte'):
    os.makedirs(pasta_saida, exist_ok=True)
    total_linhas = len(df)
    partes = (total_linhas // tamanho_lote) + 1

    for i in range(partes):
        inicio = i * tamanho_lote
        fim = inicio + tamanho_lote
        df_lote = df.iloc[inicio:fim]
        if not df_lote.empty:
            caminho_csv = os.path.join(pasta_saida, f'{nome_base}_{i+1}.csv')
            df_lote.to_csv(caminho_csv, sep=';', index=False, encoding='utf-8')
            print(f"✅ Parte {i+1} salva: {caminho_csv} ({len(df_lote)} linhas)")

"""# Caminho do arquivo original
arquivo = r'C:\Pasta_trabalho\GitLab\Python\Arquivamento Negócios Pipeline Legado (Parcial).xlsx'
pasta_saida = r'C:\Pasta_trabalho\GitLab\Python'

# ✅ Leitura com pandas
df = pd.read_excel(arquivo, sheet_name='Todos os negócios', engine='openpyxl')

# Limpeza opcional
df.dropna(how='all', inplace=True)
df.dropna(axis=1, how='all', inplace=True)

# ✅ Salva em partes
dividir_em_csvs(df, pasta_saida, tamanho_lote=5000)"""
